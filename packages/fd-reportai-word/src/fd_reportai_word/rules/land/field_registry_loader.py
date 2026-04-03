"""Load FieldGroup / FieldMetadata from an Excel workbook.

Workbook layout
---------------
- Each *sheet* represents one field group; the sheet name is the group name.
- Sheets are read in workbook order, which defines group ordering.
- Each sheet has the following columns (order-independent, matched by header):

    key            str   主键（必填，重复行会覆盖）
    kind           str   input_block / computed_field / computed_leaf / section_field（参考）
    source         str   数据来源描述
    used_by        str   消费此字段的计算字段或章节，分号分隔
    feeds_into     str   输出到哪些下游字段/章节，分号分隔
    aliases        str   别名，分号分隔
    notes          str   备注
    status         str   active / deprecated / planned（默认 active）
    priority       int   显示优先级（默认 0）
    owner          str   负责人
    relations      str   关系，格式: 关系名=字段A,字段B; 关系名2=字段C
    parent_field   str   父字段（仅结构化叶子节点填写）

Column names are case-insensitive and extra columns are ignored.
Empty / NaN cells are treated as the field default.
"""
from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    pass

from .field_registry_data import FieldGroup, FieldMetadata

_DEFAULT_XLSX_PATH = Path(__file__).resolve().parent / "land_fields.xlsx"

# --- helpers ------------------------------------------------------------------


def _split(raw: object, sep: str = ";") -> tuple[str, ...]:
    if not raw or (isinstance(raw, float) and raw != raw):  # nan check
        return ()
    parts = str(raw).split(sep)
    return tuple(p.strip() for p in parts if p.strip())


def _int_or(raw: object, default: int = 0) -> int:
    if raw is None or (isinstance(raw, float) and raw != raw):
        return default
    try:
        return int(raw)
    except (ValueError, TypeError):
        return default


def _str_or_none(raw: object) -> str | None:
    if raw is None or (isinstance(raw, float) and raw != raw):
        return None
    s = str(raw).strip()
    return s if s else None


def _parse_relations(raw: object) -> dict[str, tuple[str, ...]]:
    """Parse ``关系名=字段A,字段B; 关系名2=字段C`` into a dict."""
    if not raw or (isinstance(raw, float) and raw != raw):
        return {}
    result: dict[str, tuple[str, ...]] = {}
    for segment in str(raw).split(";"):
        segment = segment.strip()
        if not segment:
            continue
        if "=" not in segment:
            continue
        rel_name, _, values_raw = segment.partition("=")
        rel_name = rel_name.strip()
        values = tuple(v.strip() for v in values_raw.split(",") if v.strip())
        if rel_name and values:
            result[rel_name] = values
    return result


# --- public API ---------------------------------------------------------------


def load_field_registry_xlsx(
    path: Path | None = None,
) -> tuple[FieldGroup, ...]:
    """Return a tuple of :class:`FieldGroup` loaded from *path*.

    Parameters
    ----------
    path:
        Path to the ``.xlsx`` file.  Defaults to ``land_fields.xlsx`` in the
        same directory as this module.

    Returns
    -------
    tuple[FieldGroup, ...]
        Groups in workbook sheet order.  Each group's ``entries`` tuple
        preserves the row order within the sheet.
    """
    try:
        import openpyxl  # type: ignore[import]
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to load field registry from Excel. "
            "Install it with: pip install openpyxl"
        ) from exc

    xlsx_path = path or _DEFAULT_XLSX_PATH
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    groups: list[FieldGroup] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Normalise header → column index
        header_row = rows[0]
        col_index: dict[str, int] = {
            str(cell).strip().lower(): idx
            for idx, cell in enumerate(header_row)
            if cell is not None
        }

        def _get(row: tuple, col: str) -> object:
            idx = col_index.get(col)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        entries: list[FieldMetadata] = []
        for row in rows[1:]:
            key = _str_or_none(_get(row, "key"))
            if not key:
                continue  # skip blank / header continuation rows

            # notes: multi-line string → single tuple element (keep newlines)
            notes_raw = _str_or_none(_get(row, "notes"))
            notes_tuple: tuple[str, ...] = (notes_raw,) if notes_raw else ()

            entry = FieldMetadata(
                key=key,
                group=sheet_name,
                status=_str_or_none(_get(row, "status")) or "active",
                owner=_str_or_none(_get(row, "owner")),
                aliases=_split(_get(row, "aliases")),
                notes=notes_tuple,
                relations=_parse_relations(_get(row, "relations")),
                source=_str_or_none(_get(row, "source")),
                feeds_into=_split(_get(row, "feeds_into")),
                priority=_int_or(_get(row, "priority"), 0),
            )
            entries.append(entry)

        groups.append(FieldGroup(name=sheet_name, entries=tuple(entries)))

    wb.close()
    return tuple(groups)


def load_structured_computed_leaves_from_xlsx(
    path: Path | None = None,
) -> dict[str, tuple[tuple[str, str], ...]]:
    """Return ``STRUCTURED_COMPUTED_FIELD_LEAVES``-equivalent from the workbook.

    Any row with a non-empty ``parent_field`` column is treated as a structured
    leaf.  Returns a mapping of parent key → tuple of (child_key, shape) pairs
    in the row order they appear.
    """
    try:
        import openpyxl  # type: ignore[import]
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to load field registry from Excel. "
            "Install it with: pip install openpyxl"
        ) from exc

    xlsx_path = path or _DEFAULT_XLSX_PATH
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    result: dict[str, list[tuple[str, str]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header_row = rows[0]
        col_index: dict[str, int] = {
            str(cell).strip().lower(): idx
            for idx, cell in enumerate(header_row)
            if cell is not None
        }
        if "parent_field" not in col_index:
            continue

        def _get(row: tuple, col: str) -> object:
            idx = col_index.get(col)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for row in rows[1:]:
            key = _str_or_none(_get(row, "key"))
            parent = _str_or_none(_get(row, "parent_field"))
            if not key or not parent:
                continue
            # shape column is optional; default to "scalar"
            shape_raw = _str_or_none(_get(row, "shape")) or "scalar"
            result.setdefault(parent, []).append((key, shape_raw))

    wb.close()
    return {k: tuple(v) for k, v in result.items()}
